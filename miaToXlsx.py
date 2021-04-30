import sys
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment


FILENAME = './mia_demo.txt'
TEMPLATE = 'tsea28_template.xlsx'
NEW_FILENAME = 'new_tsea28.xlsx'
NAME1 = 'no name given'
NAME2 = 'no name given'

def trim_mia_file(filename: str) -> str:
    with open(filename, 'r') as mia_file:
        data = mia_file.read().replace('\n', ' ')
    return data[data.find('MyM:')+len('MyM:'):data.find('K1')]


def turn_into_dict(res: str) -> dict:
    trimmed = res.replace(': ',':').lstrip().strip()
    d = dict(item.split(':') for item in trimmed.split(' '))
    first_full_zero = list(d.keys())[list(d.values()).index('0000000')]
    only_values = dict(list(d.items())[:list(d.keys()).index(first_full_zero)])
    return only_values

def hex_to_binary(data:dict)->dict:
    binary_dict = copy(data)
    for item in binary_dict.keys():
        hex = binary_dict[item]
        out = bin(int(hex, 16))[2:].zfill(8)
        if len(out) < 25:
            zero_fillout = 25-len(out)
            out = (zero_fillout*'0') + out
        binary_dict[item] = out
    return binary_dict


def fill_xlsx(workbook,data):
    for row_index,hex in enumerate(data.values()):
        for col_index,b in enumerate(hex):
            read_sheet.cell(row=row_index+4,column=col_index+2).value = int(b)
    workbook.save(filename=NEW_FILENAME)


    
        
   
def fill_hex(workbook,data):
    for index, hex in enumerate(data.values()):
        col = 27
        for i in range(0,7):
            try:
                val = int(list(hex)[i])
            except:
                val = list(hex)[i]
            workbook['Sheet1'].cell(row=index+4,column=i+col).value = val
    workbook.save(filename=NEW_FILENAME)

##FUNKAR
def fill_adr(workbook,data)->None:
    for index,adr in enumerate(data.keys()):
        try:
            adr = int(adr)
        except:
            pass
        workbook['Sheet1'].cell(row=index+4,column=1).value = adr
    workbook.save(filename=NEW_FILENAME)

##FUNKAR
def fill_name(workbook)->None:
    workbook['Sheet1']['B1'] = NAME1
    workbook['Sheet1']['B2'] = NAME2
    workbook.save(filename=NEW_FILENAME)


if __name__ == "__main__":
    if len(sys.argv) == 3:
        NAME1 = sys.argv[1]
        NAME2 = sys.argv[2]
        FILENAME = sys.argv[3]
    

    try:
        read_from = load_workbook(TEMPLATE)
        read_sheet = read_from.active

        trimmed = trim_mia_file(FILENAME)
        data_dict = turn_into_dict(trimmed)
        binaryed_dict = hex_to_binary(data_dict)

        for cell in read_sheet['A']:
            cell.alignment = Alignment(horizontal='right')


        fill_name(read_from)
        fill_adr(read_from,data_dict)
        fill_hex(read_from,data_dict)
        fill_xlsx(read_from,binaryed_dict)
    except PermissionError as permission_error:
        print(f'You must close the .xlsx file before you rerun the script. Error: {permission_error}')
    except FileNotFoundError as file_error:
        print(f'File could not be found. Error: {file_error}')
    print('The MIA file has been converted to TSEA28 .xlsx file, you just need to fill in the comment and LIU-id.\nThe new name of the file will be new_tsea28.xlsx')
